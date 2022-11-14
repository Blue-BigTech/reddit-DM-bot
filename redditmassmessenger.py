import os
import praw
import openpyxl
import time
from dotenv import load_dotenv
from pathlib import Path


class MassMessenger:
    """
    Sends a mass message to a list of Reddit users.
    """

    def __init__(self, c_id, c_secret, user, pswd, list, message, subject):
        self.r = praw.Reddit(client_id=c_id,
                     client_secret=c_secret, password=pswd,
                     user_agent=user, username=user)
        self.list = self.get_users(list)
        self.message = self.get_contents(message)
        self.subject = self.get_contents(subject)

    def get_users(self, list):
        """
        Gets list of user's from Excel file and organizes them into a list.

        :param list: str
        :return: list[str, str]
        """
        usernames = []
        wb = openpyxl.load_workbook(list)
        ws = wb.active
        for row in range(1, ws.max_row + 1):
            usernames.append([ws['A' + str(row)].value, 'Not Sent'])
        return usernames

    def get_contents(self, location):
        """
        Gets content of text file at location.

        :param location: str
        :return: str
        """
        file = open(location, 'r')
        contents = file.read()
        if contents[-1:] == "\n":
            return contents[:len(contents)-1]
        return contents

    def run(self):
        """
        Runs main message sending code as well as runs the stats and
        creates the concluding Excel document.
        """
        self.show_stats()
        count = 0
        pause_count = 0
        print("Sending messages:")
        for user in self.list:
            if pause_count >= 100:
                print('Paused for 2 hours at: ' + str(time.ctime()))
                time.sleep(7200)
                pause_count = 0
            try:
                self.r.redditor(user[0]).message(self.subject, self.message)
            except:
                pass
            else:
                print('Mailed ' + user[0])
                self.list[count][1] = 'Sent'
                count += 1
                pause_count += 1
        self.create_excel_doc()

    def show_stats(self):
        """
        Shows basic stats on current mass message.
        
        :return: str
        """
        print('sending to ' + str(len(self.list)) + ' users.')

    def create_excel_doc(self):
        """
        Creates the final Excel document containing sending info.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Results'
        ws['A1'] = 'Username'
        ws['B1'] = 'Status'
        current_row = 2
        for user in self.list:
            ws['A' + str(current_row)] = user[0]
            ws['B' + str(current_row)] = user[1]
            current_row += 1
        wb.save('results.xlsx')

if __name__ == '__main__':
    env_path = Path('.') / '.env'
    load_dotenv(dotenv_path=env_path)

    # Setting client connection credentials
    client_id = os.getenv("client_id")
    client_secret = os.getenv("client_secret")
    username = os.getenv("reddit_username")
    password = os.getenv("reddit_password")

    # Setting data sources
    list = "sample-users.xlsx"
    message = "sample-message.txt"
    subject = "sample-subject.txt"

    m = MassMessenger(
        client_id,
        client_secret,
        username,
        password,
        list,
        message,
        subject
    )

    # Running messenger
    m.run()
